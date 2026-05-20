import os
import sys

# 프로젝트 루트 디렉토리를 sys.path에 추가
sys.path.append(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)

import asyncio
import json
from datetime import datetime

import aiohttp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from src import settings

# 커스텀 모듈 임포트
from src.api.kis_client import KisApiClient
from src.data.db_loader import load_theme_from_db
from src.data.gsheet_loader import append_stocks_to_gsheet
from src.utils.display import Colors

# =========================================================
# [설정] API 접속 정보
# =========================================================
APP_KEY = settings.KIS_API_CONFIG["app_key"]
APP_SECRET = settings.KIS_API_CONFIG["app_secret"]
ACCOUNT_ID = settings.KIS_API_CONFIG.get("account_id", "")
HTS_ID = settings.KIS_API_CONFIG.get("hts_id")

TARGET_CONDITION_NAME = settings.TARGET_CONDITION_NAME
TOKEN_FILE = str(settings.TOKEN_FILE)

print(f"[{datetime.now()}] 조건검색 (엑셀 중앙정렬 기능 추가) 시작...")

if not HTS_ID or "여기에" in HTS_ID:
    print(
        "❌ 오류: configs/kis_config.py 파일의 'hts_id'에 본인의 HTS ID를 입력해주세요!"
    )
    exit()


def safe_float(value, default=0.0):
    """문자열이나 None 값을 안전하게 float로 변환"""
    if value is None:
        return default
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------
# 헬퍼 함수: 시장 지수 등락률 파싱
# ---------------------------------------------------------
def parse_market_index_rate(data):
    if not data or data.get("rt_cd") != "0":
        return 0.0
    out1 = data.get("output1")
    if not out1:
        return 0.0
    rate_str = out1.get("bstp_nmix_prdy_ctrt") or out1.get("prdy_ctrt")
    try:
        if rate_str and float(rate_str) != 0.0:
            return float(rate_str)
        current_price = float(out1.get("bstp_nmix_prpr", "0"))
        change_amount = float(out1.get("bstp_nmix_prdy_vrss", "0"))
        prev_close = current_price - change_amount
        if prev_close != 0:
            return round((change_amount / prev_close) * 100, 2)
    except Exception:
        pass
    return 0.0


# ---------------------------------------------------------
# 상세 정보 조회 및 데이터 매핑 (비동기)
# ---------------------------------------------------------
async def fetch_single_stock(
    i,
    stock,
    total,
    sem,
    client,
    session,
    overheated_stock_codes=None,
    new_high_stock_codes=None,
    near_new_high_stock_codes=None,
    upper_limit_next_day_stock_codes=None,
    upper_limit_stock_codes=None,
):
    """단일 종목의 상세 데이터를 수집합니다.
    """
    if overheated_stock_codes is None:
        overheated_stock_codes = set()
    if new_high_stock_codes is None:
        new_high_stock_codes = set()
    if near_new_high_stock_codes is None:
        near_new_high_stock_codes = set()
    if upper_limit_next_day_stock_codes is None:
        upper_limit_next_day_stock_codes = set()
    if upper_limit_stock_codes is None:
        upper_limit_stock_codes = set()

    async with sem:
        # 요청 간의 미세한 간격을 두어 TPS 분산 (Staggering)
        await asyncio.sleep(settings.API_SLEEP_INTERVAL)
        code = stock["code"]
        name = stock["name"]

        try:
            price = int(float(stock.get("price", 0)))
            rate = float(stock.get("chgrate", 0))
        except:
            price, rate = 0, 0.0

        # KisApiClient를 사용하여 비동기로 여러 API 동시 호출
        res_detail, res_strength, res_investor, res_program = await asyncio.gather(
            client.get_current_price(session, code),
            client.get_trade_strength(session, code),
            client.get_investor_trend_estimate(session, code),
            client.get_program_net_buy(session, code),
        )

        # 실패한 API 체크
        failed_apis = []
        if res_detail.get("rt_cd") != "0":
            failed_apis.append("현재가")
        if res_strength.get("rt_cd") != "0":
            failed_apis.append("체결강도")
        if res_investor.get("rt_cd") != "0":
            failed_apis.append("투자자추정")
        if res_program.get("rt_cd") != "0":
            failed_apis.append("프로그램")

        # 데이터 파싱
        detail = res_detail.get("output") if res_detail.get("rt_cd") == "0" else None

        vol_power = 0.0
        if res_strength.get("rt_cd") == "0" and res_strength.get("output"):
            latest = res_strength["output"][0]
            vol_power = safe_float(latest.get("tday_rltv") or latest.get("ctrb_strgt"))

        frgn_qty, orgn_qty = 0, 0
        if res_investor.get("rt_cd") == "0" and res_investor.get("output2"):
            latest = res_investor["output2"][0]
            frgn_qty = int(safe_float(latest.get("frgn_fake_ntby_qty", 0)))
            orgn_qty = int(safe_float(latest.get("orgn_fake_ntby_qty", 0)))

        program_amt_won = 0.0
        if res_program.get("rt_cd") == "0" and res_program.get("output"):
            program_amt_won = safe_float(
                res_program["output"][0].get("whol_smtn_ntby_tr_pbmn", 0)
            )
        if detail:
            try:
                # 가격 데이터 추출
                close_price = int(safe_float(detail.get("stck_prpr"), price))
                open_price = int(safe_float(detail.get("stck_oprc"), 0))
                high_price = int(safe_float(detail.get("stck_hgpr"), 0))
                low_price = int(safe_float(detail.get("stck_lwpr"), 0))
                vol_acml = int(safe_float(detail.get("acml_vol"), 0))

                rate = safe_float(detail.get("prdy_ctrt"), rate)

                # [New] 단기과열 종목 확인 (HTS 조건검색 기준)
                is_overheated = code in overheated_stock_codes
                if is_overheated:
                    print(f"\n  🔥 {name}: 단기과열 종목 (HTS) → 차트통과=0 설정 예정")

                # 전일 종가 계산
                if rate != 0:
                    prev_close_price = int(close_price / (1 + rate / 100))
                else:
                    prev_close_price = close_price

                price = close_price
                shares = safe_float(detail.get("lstn_stcn"), 0)
                raw_market = str(detail.get("rprs_mrkt_kor_name", "")).upper()
                market_name = (
                    "KOSPI"
                    if "KOSPI" in raw_market or "유가" in raw_market
                    else "KOSDAQ" if "KOSDAQ" in raw_market else raw_market
                )
                raw_mkt_cap = safe_float(detail.get("hts_avls")) * 100_000_000
                if raw_mkt_cap == 0:
                    raw_mkt_cap = shares * price
                mkt_cap_eok = round(raw_mkt_cap / 100_000_000, 2)
                trade_amt_eok = round(
                    safe_float(detail.get("acml_tr_pbmn")) / 100_000_000, 2
                )
            except Exception as e:
                print(f"\n {Colors.RED}⚠️ [{name}] 파싱 에러: {e}{Colors.RESET}")

        frgn_net_eok = round((frgn_qty * price) / 100_000_000, 2)
        orgn_net_eok = round((orgn_qty * price) / 100_000_000, 2)
        program_net_eok = round(program_amt_won / 100_000_000, 2)

        ema_multi_res = {}
        chart_pass = 1
        scenario = ""
        data_count = 0
        sma_value, sma_success = 0, False
        sma60_value, sma60_success = 0, False

        # EMA & SMA 계산 (중복 API 호출 방지 통합 버전)
        try:
            from src.api.kis_client import calculate_all_moving_averages

            (
                ema_multi_res,
                (ema20_val, ema_success, data_count),
                (sma60_val, sma60_ok),
                (sma120_val, sma120_ok),
            ) = await calculate_all_moving_averages(code, session=session)

            sma_value, sma_success = float(sma120_val), sma120_ok
            sma60_value, sma60_success = float(sma60_val), sma60_ok

        except Exception:
            pass

        # === 7가지 필터링 및 시나리오 분류 (우선순위 역순 적용) ===
        ema5 = ema_multi_res.get(5, 0)
        ema10 = ema_multi_res.get(10, 0)
        ema20 = ema_multi_res.get(20, 0)

        # 1. 단기과열 필터링
        if code in overheated_stock_codes:
            chart_pass = 0

        # 2. 상장일수 부족 필터링
        if data_count < settings.EMA_PERIOD:
            chart_pass = 0

        # === 시나리오 할당 (우선순위: 높음 → 낮음) ===
        # 우선순위: 상따(상한가) > 상한가 다음날 > 상승형 음봉 > 120 돌파 > 신고가 근접 > 신고가

        # 1. [최최우선] 상따 (상한가 조건검색)
        if code in upper_limit_stock_codes:
            scenario = "상따"
            # 상한가는 차트 필터링 없이 통과

        # 2. 상한가 다음날
        elif code in upper_limit_next_day_stock_codes:
            scenario = "상한가 다음날"
            # SMA 120 아래면 차트 필터링
            if sma_success and sma_value > 0 and close_price < sma_value:
                chart_pass = 0

        # 3. 상승형 음봉 (종가 < 시가 이지만 등락률 > 0)
        elif rate > 0 and close_price < open_price:
            scenario = "상승형 음봉"
            # 이평선 아래면 차트 필터링
            if (
                (ema5 > 0 and close_price < ema5)
                or (ema10 > 0 and close_price < ema10)
                or (ema20 > 0 and close_price < ema20)
                or (sma60_success and close_price < sma60_value)
                or (sma_success and close_price < sma_value)
            ):
                chart_pass = 0

        # 4. 120 돌파
        elif (
            sma_success
            and sma_value > 0
            and prev_close_price < sma_value <= close_price
        ):
            scenario = "120 돌파"
            # 단기 이평선 아래면 차트 필터링
            if (
                (ema5 > 0 and close_price < ema5)
                or (ema10 > 0 and close_price < ema10)
                or (ema20 > 0 and close_price < ema20)
                or (sma60_success and close_price < sma60_value)
            ):
                chart_pass = 0

        # 5. 신고가 근접 (우선순위 상승)
        elif code in near_new_high_stock_codes:
            scenario = "신고가 근접"
            # 이평선 아래면 차트 필터링
            if (
                (ema5 > 0 and close_price < ema5)
                or (ema10 > 0 and close_price < ema10)
                or (ema20 > 0 and close_price < ema20)
                or (sma60_success and close_price < sma60_value)
                or (sma_success and close_price < sma_value)
            ):
                chart_pass = 0

        # 6. 신고가
        elif code in new_high_stock_codes:
            scenario = "신고가"
            # 이평선 아래면 차트 필터링
            if (
                (ema5 > 0 and close_price < ema5)
                or (ema10 > 0 and close_price < ema10)
                or (ema20 > 0 and close_price < ema20)
                or (sma60_success and close_price < sma60_value)
                or (sma_success and close_price < sma_value)
            ):
                chart_pass = 0

        print(
            f"\r {Colors.YELLOW}🔍 데이터 수집 중... ({i+1}/{total}){Colors.RESET}",
            end="",
            flush=True,
        )

        return {
            "종목명": name,
            "종목코드": code,
            "시가": open_price,
            "고가": high_price,
            "저가": low_price,
            "종가": close_price,
            "전일종가": prev_close_price,
            "시장구분": market_name,
            "시가총액(억)": mkt_cap_eok,
            "거래대금(억)": trade_amt_eok,
            "체결강도": vol_power,
            "등락률": rate,
            "순위": 0,
            "기관_순매수(억)": orgn_net_eok,
            "외국인_순매수(억)": frgn_net_eok,
            "프로그램_순매수(억)": program_net_eok,
            "(차트통과)": chart_pass,
            "(시나리오)": scenario,
            "(상장일수)": data_count,
            "(거래량)": vol_acml,
            "(sma60)": round(sma60_value, 2) if sma60_success else 0,
            "(sma120)": round(sma_value, 2) if sma_success else 0,
        }, failed_apis


async def fetch_all_stock_data(
    stock_list,
    client,
    session,
    overheated_stock_codes=None,
    new_high_stock_codes=None,
    near_new_high_stock_codes=None,
    upper_limit_next_day_stock_codes=None,
    upper_limit_stock_codes=None,
):
    """모든 종목의 상세 데이터를 수집합니다.
    """
    if overheated_stock_codes is None:
        overheated_stock_codes = set()
    if new_high_stock_codes is None:
        new_high_stock_codes = set()
    if near_new_high_stock_codes is None:
        near_new_high_stock_codes = set()
    if upper_limit_next_day_stock_codes is None:
        upper_limit_next_day_stock_codes = set()
    if upper_limit_stock_codes is None:
        upper_limit_stock_codes = set()

    sem = asyncio.Semaphore(settings.API_SEMAPHORE_LIMIT)
    total = len(stock_list)
    tasks = [
        fetch_single_stock(
            i,
            stock,
            total,
            sem,
            client,
            session,
            overheated_stock_codes,
            new_high_stock_codes,
            near_new_high_stock_codes,
            upper_limit_next_day_stock_codes,
            upper_limit_stock_codes,
        )
        for i, stock in enumerate(stock_list)
    ]
    all_res = await asyncio.gather(*tasks)

    results = [r for r, f in all_res]
    failed_info = [
        (stock_list[i]["name"], stock_list[i]["code"], f)
        for i, (r, f) in enumerate(all_res)
        if f
    ]

    print(f"\n{Colors.GREEN}✅ 데이터 수집 완료{Colors.RESET}")
    return results, failed_info


async def main():
    from aiohttp.resolver import ThreadedResolver

    # aiohttp 세션 설정 강화 (네트워크 안정성 향상 + DNS 해결)
    timeout = aiohttp.ClientTimeout(
        total=60,  # 전체 요청 타임아웃
        connect=10,  # 연결 타임아웃
        sock_read=30,  # 소켓 읽기 타임아웃
    )
    connector = aiohttp.TCPConnector(
        limit=20,  # 최대 동시 연결 수
        ttl_dns_cache=300,  # DNS 캐시 TTL (5분)
        force_close=False,  # Keep-Alive 유지
        resolver=ThreadedResolver(),  # [Fix] Windows aiodns 이슈 방지용 표준 리졸버 사용
    )

    async with aiohttp.ClientSession(timeout=timeout, connector=connector) as session:
        # 1. 클라이언트 초기화 및 토큰 확보
        client = KisApiClient(
            APP_KEY, APP_SECRET, ACCOUNT_ID, HTS_ID, token_file=TOKEN_FILE
        )
        await client.ensure_token(session)

        # 2. 시장 지수 조회
        res_kospi = await client.get_market_index_rate(session, "0001")
        res_kosdaq = await client.get_market_index_rate(session, "1001")

        kospi_rate = parse_market_index_rate(res_kospi)
        kosdaq_rate = parse_market_index_rate(res_kosdaq)

        # 3. 조건 목록 및 대상 조건 찾기
        res_cond_list = await client.get_condition_list(session)
        if res_cond_list.get("rt_cd") != "0":
            print(
                f"{Colors.RED}조건식 목록 조회 실패: "
                f"rt_cd={res_cond_list.get('rt_cd')}, "
                f"msg={res_cond_list.get('msg1', 'N/A')}{Colors.RESET}"
            )
            return
        my_conditions = res_cond_list.get("output2", [])

        if not my_conditions:
            print(
                f"{Colors.YELLOW}⚠ 저장된 조건이 없습니다. "
                f"(HTS ID: {HTS_ID}, output2 비어 있음){Colors.RESET}"
            )
            return
        target_cond = next(
            (c for c in my_conditions if TARGET_CONDITION_NAME in c["condition_nm"]),
            None,
        )
        if not target_cond:
            print(
                f"{Colors.RED}❌ '{TARGET_CONDITION_NAME}' 조건을 찾을 수 없습니다.{Colors.RESET}"
            )
            return

        print(
            f"\n>> {Colors.BOLD}[{target_cond['condition_nm']}]{Colors.RESET} 검색 시작..."
        )

        # 4. 조건검색 결과 조회 (종가매매)
        res_cond_res = await client.get_condition_result(session, target_cond["seq"])
        if res_cond_res.get("rt_cd") != "0":
            print(f"{Colors.RED}❌ 검색 실패: {res_cond_res.get('msg1')}{Colors.RESET}")
            return

        stock_list = res_cond_res.get("output2", [])
        print(
            f"{Colors.GREEN}✅ 검색 결과: 총 {len(stock_list)} 종목 포착!{Colors.RESET}"
        )

        # 4-1. [New] 단기과열 조건검색 결과 조회 (예고 + 본지정)
        overheated_stock_codes = set()  # 단기과열 종목 코드 Set

        # [FIX] 정확한 일치로 변경
        overheated_cond = next(
            (
                c
                for c in my_conditions
                if c["condition_nm"] == settings.OVERHEATED_CONDITION_NAME
            ),
            None,
        )

        if overheated_cond:
            print(
                f"\n>> {Colors.YELLOW}[{overheated_cond['condition_nm']}]{Colors.RESET} 검색 중..."
            )
            res_overheated = await client.get_condition_result(
                session, overheated_cond["seq"]
            )

            if res_overheated.get("rt_cd") == "0":
                overheated_list = res_overheated.get("output2", [])
                overheated_stock_codes = {
                    stock.get("code") for stock in overheated_list
                }
                print(
                    f"{Colors.YELLOW}🔥 단기과열 종목: {len(overheated_stock_codes)}개 포착{Colors.RESET}"
                )

                # 교집합 확인 (종가매매 종목 중 단기과열 종목)
                main_stock_codes = {stock.get("code") for stock in stock_list}
                overlap = main_stock_codes & overheated_stock_codes
                if overlap:
                    print(
                        f"{Colors.YELLOW}   → 종가매매와 중복: {len(overlap)}개 종목{Colors.RESET}"
                    )
            else:
                print(
                    f"{Colors.YELLOW}⚠️  단기과열 조건검색 실패 (무시하고 진행){Colors.RESET}"
                )
        else:
            print(
                f"\n{Colors.YELLOW}⚠️  '{settings.OVERHEATED_CONDITION_NAME}' 조건을 찾을 수 없습니다.{Colors.RESET}"
            )
            print(
                f"   → HTS에서 '{settings.OVERHEATED_CONDITION_NAME}' 조건을 생성해주세요."
            )
            print("   → 단기과열 필터링이 비활성화됩니다.")

        # 4-2. [NEW] 신고가 조건검색 결과 조회
        new_high_stock_codes = set()
        new_high_cond_name = settings.NEW_HIGH_CONDITION_NAME
        print(
            f"\n{Colors.CYAN}[디버그] '{new_high_cond_name}' 조건검색 조회 중...{Colors.RESET}"
        )
        # [FIX] 정확한 일치로 변경 ("신고가" in "신고가 근접" 방지)
        new_high_cond = next(
            (c for c in my_conditions if c["condition_nm"] == new_high_cond_name), None
        )

        if new_high_cond:
            print(
                f"  ✓ 조건 발견: {new_high_cond['condition_nm']} (seq: {new_high_cond['seq']})"
            )
            res_new_high = await client.get_condition_result(
                session, new_high_cond["seq"]
            )
            print(
                f"  API 응답: rt_cd={res_new_high.get('rt_cd')}, msg={res_new_high.get('msg1', 'N/A')}"
            )

            if res_new_high.get("rt_cd") == "0":
                new_high_list = res_new_high.get("output2", [])
                new_high_stock_codes = {stock.get("code") for stock in new_high_list}
                print(
                    f"{Colors.GREEN}🚀 [신고가]: {len(new_high_stock_codes)}개 종목 포착{Colors.RESET}"
                )
                if new_high_stock_codes:
                    sample_names = [
                        f"{s.get('name', 'N/A')}({s.get('code', 'N/A')})"
                        for s in new_high_list[:3]
                    ]
                    print(f"  샘플: {', '.join(sample_names)}")
            else:
                print(
                    f"{Colors.YELLOW}⚠️  [신고가] API 호출 실패 (rt_cd={res_new_high.get('rt_cd')}){Colors.RESET}"
                )
        else:
            print(
                f"{Colors.YELLOW}⚠️  '{new_high_cond_name}' 조건을 찾을 수 없습니다.{Colors.RESET}"
            )
            print("   → HTS에 등록된 조건검색 목록 확인 필요")
            print(
                f"   → 등록된 조건 ({len(my_conditions)}개): {[c['condition_nm'] for c in my_conditions[:5]]}"
            )

        # 4-3. [NEW] 신고가 근접 조건검색 결과 조회
        near_new_high_stock_codes = set()
        near_new_high_cond_name = settings.NEAR_NEW_HIGH_CONDITION_NAME
        print(
            f"\n{Colors.CYAN}[디버그] '{near_new_high_cond_name}' 조건검색 조회 중...{Colors.RESET}"
        )
        # [FIX] 정확한 일치로 변경
        near_new_high_cond = next(
            (c for c in my_conditions if c["condition_nm"] == near_new_high_cond_name),
            None,
        )

        if near_new_high_cond:
            print(
                f"  ✓ 조건 발견: {near_new_high_cond['condition_nm']} (seq: {near_new_high_cond['seq']})"
            )
            res_near_new_high = await client.get_condition_result(
                session, near_new_high_cond["seq"]
            )
            print(
                f"  API 응답: rt_cd={res_near_new_high.get('rt_cd')}, msg={res_near_new_high.get('msg1', 'N/A')}"
            )

            if res_near_new_high.get("rt_cd") == "0":
                near_new_high_list = res_near_new_high.get("output2", [])
                near_new_high_stock_codes = {
                    stock.get("code") for stock in near_new_high_list
                }
                print(
                    f"{Colors.YELLOW}📈 [신고가 근접]: {len(near_new_high_stock_codes)}개 종목 포착{Colors.RESET}"
                )
                if near_new_high_stock_codes:
                    sample_names = [
                        f"{s.get('name', 'N/A')}({s.get('code', 'N/A')})"
                        for s in near_new_high_list[:3]
                    ]
                    print(f"  샘플: {', '.join(sample_names)}")
            else:
                print(
                    f"{Colors.YELLOW}⚠️  [신고가 근접] API 호출 실패 (rt_cd={res_near_new_high.get('rt_cd')}){Colors.RESET}"
                )
        else:
            print(
                f"{Colors.YELLOW}⚠️  '{near_new_high_cond_name}' 조건을 찾을 수 없습니다.{Colors.RESET}"
            )
            print("   → HTS에 등록된 조건검색 목록 확인 필요")

        # 4-4. [NEW] 상한가 다음날 조건검색 결과 조회
        upper_limit_next_day_stock_codes = set()
        upper_limit_cond_name = settings.UPPER_LIMIT_NEXT_DAY_CONDITION_NAME
        # [FIX] 정확한 일치로 변경
        upper_limit_cond = next(
            (c for c in my_conditions if c["condition_nm"] == upper_limit_cond_name),
            None,
        )
        if upper_limit_cond:
            res_upper_limit = await client.get_condition_result(
                session, upper_limit_cond["seq"]
            )
            if res_upper_limit.get("rt_cd") == "0":
                upper_limit_list = res_upper_limit.get("output2", [])
                upper_limit_next_day_stock_codes = {
                    stock.get("code") for stock in upper_limit_list
                }
                print(
                    f"{Colors.MAGENTA}⬆️ [상한가 다음날]: {len(upper_limit_next_day_stock_codes)}개 종목 포착{Colors.RESET}"
                )

        # 4-5. [NEW] 상한가 조건검색 결과 조회
        upper_limit_stock_codes = set()
        upper_limit_cond_name = settings.UPPER_LIMIT_CONDITION_NAME
        upper_limit_cond = next(
            (c for c in my_conditions if c["condition_nm"] == upper_limit_cond_name),
            None,
        )
        if upper_limit_cond:
            res_upper_limit_only = await client.get_condition_result(
                session, upper_limit_cond["seq"]
            )
            if res_upper_limit_only.get("rt_cd") == "0":
                upper_limit_only_list = res_upper_limit_only.get("output2", [])
                upper_limit_stock_codes = {
                    stock.get("code") for stock in upper_limit_only_list
                }
                print(
                    f"{Colors.RED}🔺 [상한가]: {len(upper_limit_stock_codes)}개 종목 포착 (일반분석 제외){Colors.RESET}"
                )

        # 5. 상세 데이터 수집 (HTS 조건검색 결과 전달)
        results, failed_info = await fetch_all_stock_data(
            stock_list,
            client,
            session,
            overheated_stock_codes,
            new_high_stock_codes,
            near_new_high_stock_codes,
            upper_limit_next_day_stock_codes,
            upper_limit_stock_codes,
        )
        if results:
            # 데이터 저장 및 후처리 로직 (기존과 동일)
            clean_name = (
                target_cond["condition_nm"].replace("/", "_").replace("\\", "_")
            )
            save_path = str(settings.DATA_DIR / f"condition_{clean_name}.xlsx")

            df = pd.DataFrame(results)
            df = df.sort_values(by="거래대금(억)", ascending=False)
            df["순위"] = range(1, len(df) + 1)

            # 캐시 로직
            cache_path = str(settings.CHART_PASS_CACHE_FILE)
            today_str = datetime.now().strftime("%Y-%m-%d")
            daily_memory = {}

            if os.path.exists(cache_path):
                try:
                    with open(cache_path, encoding="utf-8") as f:
                        cache_data = json.load(f)
                        if cache_data.get("date") == today_str:
                            daily_memory = cache_data.get("data", {})
                except:
                    pass

            # 신규 종목 감지 및 캐시 업데이트
            current_codes = set(df["종목코드"].unique())
            cached_codes = set(daily_memory.keys())
            new_codes = current_codes - cached_codes

            if new_codes:
                print(
                    f"\n{Colors.MAGENTA}✨ [신규 종목 포착] {len(new_codes)}개 종목이 새로 발견되었습니다.{Colors.RESET}"
                )
                for code in new_codes:
                    # 종목코드에 해당하는 종목명을 찾아 함께 출력
                    name = df[df["종목코드"] == code]["종목명"].iloc[0]
                    print(f"   👉 {name} ({code})")
                    daily_memory[code] = 1

            # 기존 엑셀 수정사항 반영
            if os.path.exists(save_path):
                try:
                    if (
                        datetime.fromtimestamp(os.path.getmtime(save_path)).date()
                        == datetime.now().date()
                    ):
                        old_df = pd.read_excel(save_path, engine="openpyxl")
                        if (
                            "종목코드" in old_df.columns
                            and "(차트통과)" in old_df.columns
                        ):
                            old_df["종목코드"] = (
                                old_df["종목코드"].astype(str).str.zfill(6)
                            )
                            daily_memory.update(
                                dict(zip(old_df["종목코드"], old_df["(차트통과)"]))
                            )
                except:
                    pass

            # 캔들 몸통 비율 계산 및 필터링
            def calculate_candle_body_ratio(row):
                """캔들의 몸통 비율을 계산 (몸통 크기 / 전체 크기)"""
                open_price = row["시가"]
                high_price = row["고가"]
                low_price = row["저가"]
                close_price = row["종가"]

                # 가격 데이터가 유효한지 확인
                if high_price == 0 or low_price == 0 or high_price == low_price:
                    return 0.0

                body_size = abs(close_price - open_price)  # 몸통 크기
                total_size = high_price - low_price  # 전체 크기 (고가 - 저가)

                if total_size == 0:
                    return 0.0

                return body_size / total_size

            # 각 종목의 캔들 몸통 비율 계산
            df["캔들몸통비율"] = df.apply(calculate_candle_body_ratio, axis=1)

            # [New] 시가 갭 상승 비율 계산
            def calculate_gap_ratio(row):
                """시가 갭 비율 계산 (시가 - 전일종가) / 전일종가"""
                prev_close = row["전일종가"]
                open_price = row["시가"]

                if prev_close == 0:
                    return 0.0

                return (open_price - prev_close) / prev_close

            df["시가갭비율"] = df.apply(calculate_gap_ratio, axis=1)

            # [수정] 차트통과 값은 이미 fetch_single_stock에서 SMA/EMA/단기과열 기준으로 설정됨
            # 추가로 캔들 몸통 비율 필터링 적용 (단, 시가 갭 상승이 높으면 예외)

            # SMA/EMA/단기과열로 이미 필터된 종목 확인
            sma_filtered_count = len(df[df["(차트통과)"] == 0])

            # 캔들 몸통이 약하지만 시가 갭 상승이 큰 경우 예외 처리
            # 조건: 캔들몸통비율 < 50% AND 시가갭비율 < 5%인 경우만 필터링
            weak_candle_mask = (
                df["캔들몸통비율"] < settings.CANDLE_BODY_RATIO_THRESHOLD
            ) & (
                df["시가갭비율"] < settings.GAP_UP_THRESHOLD
            )  # 갭 상승이 작을 때만 필터

            # [수정] 상한가 다음날 시나리오는 SMA 120 위라면 몸통 비율 예외 적용
            if "(시나리오)" in df.columns and "(sma120)" in df.columns:
                upper_limit_exception = (df["(시나리오)"] == "상한가 다음날") & (
                    df["종가"] >= df["(sma120)"]
                )
                weak_candle_mask = weak_candle_mask & (~upper_limit_exception)

            # 차트통과가 1인 종목 중에서 약한 캔들 필터링
            df.loc[weak_candle_mask & (df["(차트통과)"] == 1), "(차트통과)"] = 0

            # 최종 필터링 결과 출력
            total_filtered = len(df[df["(차트통과)"] == 0])
            candle_filtered_count = total_filtered - sma_filtered_count

            print(f"\n{Colors.CYAN}{'='*60}{Colors.RESET}")
            print(f"{Colors.CYAN}[차트 필터링 결과]{Colors.RESET}")

            # 단기과열 필터링 종목 (이미 앞에서 출력되었지만 요약)
            if overheated_stock_codes:
                overheated_in_list = df[df["종목코드"].isin(overheated_stock_codes)]
                if not overheated_in_list.empty:
                    print(
                        f"\n  🔥 {Colors.YELLOW}단기과열: {len(overheated_in_list)}개{Colors.RESET}"
                    )
                    for _, row in overheated_in_list.iterrows():
                        print(f"     - {row['종목명']} ({row['종목코드']})")

            # SMA 120 필터링 종목 (신규종목 포함)
            if sma_filtered_count > 0:
                print(
                    f"\n  📉 {Colors.RED}SMA 120 필터: {sma_filtered_count}개{Colors.RESET}"
                )
                # 실제로는 단기과열이 아니면서 차트통과=0인 종목 중 캔들 몸통은 괜찮은 것들
                sma_filtered = df[
                    (df["(차트통과)"] == 0)
                    & (~df["종목코드"].isin(overheated_stock_codes))
                    & (df["캔들몸통비율"] >= 0.5)
                ]
                for _, row in sma_filtered.head(10).iterrows():  # 최대 10개만 표시
                    print(f"     - {row['종목명']} ({row['종목코드']})")
                if len(sma_filtered) > 10:
                    print(f"     ... 외 {len(sma_filtered) - 10}개")

            # 캔들 몸통 필터링 종목 (갭 상승 예외 제외)
            if candle_filtered_count > 0:
                print(
                    f"\n  📊 {Colors.YELLOW}캔들 몸통 필터: {candle_filtered_count}개{Colors.RESET}"
                )
                candle_filtered = df[
                    (df["(차트통과)"] == 0)
                    & (~df["종목코드"].isin(overheated_stock_codes))
                    & (df["캔들몸통비율"] < settings.CANDLE_BODY_RATIO_THRESHOLD)
                    & (
                        df["시가갭비율"] < settings.GAP_UP_THRESHOLD
                    )  # 갭 상승 작은 것만
                ]
                for _, row in candle_filtered.head(10).iterrows():
                    print(
                        f"     - {row['종목명']} (몸통: {row['캔들몸통비율']:.1%}, 갭: {row['시가갭비율']:+.1%})"
                    )
                if len(candle_filtered) > 10:
                    print(f"     ... 외 {len(candle_filtered) - 10}개")

            # [New] 갭 상승 예외로 통과한 종목 (캔들 몸통은 약하지만 갭 상승이 큼)
            gap_exception = df[
                (df["(차트통과)"] == 1)
                & (df["캔들몸통비율"] < settings.CANDLE_BODY_RATIO_THRESHOLD)
                & (df["시가갭비율"] >= settings.GAP_UP_THRESHOLD)
            ]
            if not gap_exception.empty:
                print(
                    f"\n  🚀 {Colors.GREEN}갭 상승 예외 통과: {len(gap_exception)}개{Colors.RESET}"
                )
                print(
                    f"     (캔들 몸통 < {settings.CANDLE_BODY_RATIO_THRESHOLD:.0%} 이지만 갭 상승 >= {settings.GAP_UP_THRESHOLD:.0%})"
                )
                for _, row in gap_exception.head(10).iterrows():
                    print(
                        f"     - {row['종목명']} (몸통: {row['캔들몸통비율']:.1%}, 갭: {row['시가갭비율']:+.1%})"
                    )
                if len(gap_exception) > 10:
                    print(f"     ... 외 {len(gap_exception) - 10}개")

            # 요약
            print(f"\n  {Colors.BOLD}[요약]{Colors.RESET}")
            print(f"  총 필터: {Colors.RED}{total_filtered}{Colors.RESET}개 종목")
            print(
                f"  통과: {Colors.GREEN}{len(df) - total_filtered}{Colors.RESET}개 종목"
            )
            print(f"{Colors.CYAN}{'='*60}{Colors.RESET}\n")

            # 캔들몸통비율 컬럼은 임시 분석용이므로 최종 저장 시 제외

            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump({"date": today_str, "data": daily_memory}, f, indent=2)

            # 추가 지표 계산
            total_count = len(df)
            df["전체종목수"] = total_count
            df["평균거래대금(억)"] = round(df["거래대금(억)"].mean(), 2)
            df["KOSPI등락률"] = kospi_rate
            df["KOSDAQ등락률"] = kosdaq_rate

            # [New] V-KOSPI & V-KOSDAQ 계산 및 추가
            try:
                from src.api.kis_client import fetch_index_and_calculate_volatility

                # V-KOSPI (KOSPI 200: 1028)
                vkospi_val, vkospi_chg = await fetch_index_and_calculate_volatility(
                    "1028", session=session
                )
                df["(v-kospi)"] = round(vkospi_val, 2)

                # V-KOSDAQ (KOSDAQ 150: 2203)
                vkosdaq_val, vkosdaq_chg = await fetch_index_and_calculate_volatility(
                    "2203", session=session
                )
                df["(v-kosdaq)"] = round(vkosdaq_val, 2)

            except Exception:
                df["(v-kospi)"] = 0.0
                df["(v-kosdaq)"] = 0.0

            # 컬럼명 리네임 및 순서 재배치를 적용할 임시 데이터프레임 생성
            rename_map = {
                "종목코드": "(종목코드)",
                "시가": "(시가)",
                "고가": "(고가)",
                "저가": "(저가)",
                "종가": "(종가)",
                "전일종가": "(전일종가)",
                "시가총액(억)": "(시가총액, 억)",
                "거래대금(억)": "(거래대금, 억)",
                "등락률": "(등락률)",
                "순위": "(선정 순위)",
                "기관_순매수(억)": "(기관_순매수)",
                "외국인_순매수(억)": "(외국인_순매수)",
                "프로그램_순매수(억)": "(프로그램_순매수)",
                "체결강도": "(체결강도)",
                "시장구분": "(시장구분)",
                "전체종목수": "(총 종목 수)",
                "평균거래대금(억)": "(평균 거래대금)",
                "KOSPI등락률": "(kospi, %)",
                "KOSDAQ등락률": "(kosdaq, %)",
            }
            
            df_excel = df.rename(columns=rename_map)
            
            cols_order = [
                "(차트통과)",
                "(시나리오)",
                "(상장일수)",
                "종목명",
                "(종목코드)",
                "(시가)",
                "(고가)",
                "(저가)",
                "(종가)",
                "(전일종가)",
                "(시가총액, 억)",
                "(거래대금, 억)",
                "(등락률)",
                "(선정 순위)",
                "(기관_순매수)",
                "(외국인_순매수)",
                "(프로그램_순매수)",
                "(체결강도)",
                "(시장구분)",
                "(총 종목 수)",
                "(평균 거래대금)",
                "(kospi, %)",
                "(kosdaq, %)",
                "(v-kospi)",
                "(v-kosdaq)",
                "(거래량)",
            ]
            df_excel[cols_order].to_excel(save_path, index=False)

            # 엑셀 서식 적용
            try:
                wb = load_workbook(save_path)
                ws = wb.active
                center_align = Alignment(horizontal="center", vertical="center")
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = center_align
                wb.save(save_path)
            except:
                pass

            # 수집 결과 요약 출력
            success_count = len(results) - len(failed_info)

            # 테마 미매칭 종목 확인 및 구글 시트 자동 등록
            theme_map = load_theme_from_db()
            df["테마"] = df["종목코드"].map(theme_map)
            no_theme_df = df[df["테마"].isna() | (df["테마"] == "")]

            if not no_theme_df.empty:
                no_theme_list = no_theme_df[["종목코드", "종목명"]].to_dict("records")
                print(
                    f"\n{Colors.BOLD}⚠️ [테마 미매칭] {Colors.YELLOW}{len(no_theme_list)}{Colors.RESET} 종목 발견"
                )
                print(
                    f"   👉 대상 종목: {', '.join([s['종목명'] for s in no_theme_list])}"
                )
                print("   [진행] 미매칭 종목을 구글 시트에 등록 중...")

                key_path = str(settings.GOOGLE_KEY_PATH)
                GOOGLE_SHEET_NAME = settings.GOOGLE_SHEET_NAME
                THEME_WORKSHEET_NAME = settings.THEME_WORKSHEET_NAME
                append_stocks_to_gsheet(
                    key_path, GOOGLE_SHEET_NAME, THEME_WORKSHEET_NAME, no_theme_list
                )
            print(f"\n{Colors.BOLD}� [데이터 수집 요약]{Colors.RESET}")
            print(f"   ✅ 성공: {Colors.GREEN}{success_count}{Colors.RESET} 종목")
            if failed_info:
                print(f"   ❌ 실패: {Colors.RED}{len(failed_info)}{Colors.RESET} 종목")
                for name, code, apis in failed_info:
                    print(
                        f"      - {name} ({code}) [실패항목: {Colors.YELLOW}{', '.join(apis)}{Colors.RESET}]"
                    )

            print(f"\n{Colors.GREEN}📂 엑셀 저장 완료: {save_path}{Colors.RESET}")
        else:
            print(f"{Colors.YELLOW}⚠ 검색된 종목이 없습니다.{Colors.RESET}")


if __name__ == "__main__":
    asyncio.run(main())
